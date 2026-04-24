"""Load Model IQ exports from JSON, CSV, XLSX, YAML/YML into normalized records."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable

from ..models.domain import ModelIQRecord, Urgency
from ..scanners.deployment_type import classify_text as classify_deployment_type
from ..utils.dates import parse_date
from ..utils.text import normalize_key


# Synonyms → canonical field name on ModelIQRecord.
FIELD_ALIASES: dict[str, str] = {
    "deployment_name": "deployment_name",
    "deployment": "deployment_name",
    "deployment_id": "deployment_name",
    "name": "deployment_name",
    "current_model": "current_model",
    "model": "current_model",
    "model_name": "current_model",
    "current_version": "current_version",
    "version": "current_version",
    "model_version": "current_version",
    "recommended_replacement": "recommended_replacement",
    "replacement": "recommended_replacement",
    "target_model": "recommended_replacement",
    "recommended_model": "recommended_replacement",
    "retirement_date": "retirement_date",
    "retire_by": "retirement_date",
    "deprecation_date": "retirement_date",
    "eol_date": "retirement_date",
    "region": "region",
    "location": "region",
    "subscription": "subscription",
    "subscription_id": "subscription",
    "sku": "sku",
    "sku_name": "sku",
    "offering": "sku",
    "deployment_type": "deployment_type",
    "deployment_sku": "deployment_type",
    "tier": "deployment_type",
    "capacity": "capacity",
    "ptu_count": "capacity",
    "ptu": "capacity",
    "provisioned_capacity": "capacity",
    "environment": "environment",
    "env": "environment",
    "urgency": "urgency",
    "urgency_bucket": "urgency",
    "priority": "urgency",
    "notes": "notes",
    "advisory": "notes",
    "advisory_text": "notes",
    "comments": "notes",
}


_URGENCY_MAP = {
    "immediate": Urgency.IMMEDIATE,
    "critical": Urgency.IMMEDIATE,
    "urgent": Urgency.IMMEDIATE,
    "high": Urgency.HIGH,
    "medium": Urgency.MEDIUM,
    "med": Urgency.MEDIUM,
    "moderate": Urgency.MEDIUM,
    "low": Urgency.LOW,
    "info": Urgency.LOW,
    "informational": Urgency.LOW,
    "": Urgency.UNKNOWN,
}


def _coerce_urgency(value: Any) -> Urgency:
    if value is None:
        return Urgency.UNKNOWN
    return _URGENCY_MAP.get(str(value).strip().lower(), Urgency.UNKNOWN)


def _row_to_record(row: dict[str, Any]) -> ModelIQRecord:
    normalized: dict[str, Any] = {}
    raw: dict[str, Any] = {}
    for k, v in row.items():
        if k is None:
            continue
        raw[str(k)] = v
        canonical = FIELD_ALIASES.get(normalize_key(str(k)))
        if canonical:
            normalized[canonical] = v

    # Deployment type: explicit field wins, otherwise infer from SKU/offering text.
    dtype_raw = normalized.get("deployment_type") or normalized.get("sku")
    deployment_type = classify_deployment_type(str(dtype_raw) if dtype_raw is not None else None)

    capacity_val = normalized.get("capacity")
    capacity: int | None
    try:
        capacity = int(str(capacity_val).strip()) if capacity_val not in (None, "") else None
    except (TypeError, ValueError):
        capacity = None

    return ModelIQRecord(
        deployment_name=(str(normalized["deployment_name"]).strip() if normalized.get("deployment_name") else None),
        current_model=(str(normalized["current_model"]).strip() if normalized.get("current_model") else None),
        current_version=(str(normalized["current_version"]).strip() if normalized.get("current_version") else None),
        recommended_replacement=(str(normalized["recommended_replacement"]).strip()
                                 if normalized.get("recommended_replacement") else None),
        retirement_date=parse_date(normalized.get("retirement_date")),
        region=(str(normalized["region"]).strip() if normalized.get("region") else None),
        subscription=(str(normalized["subscription"]).strip() if normalized.get("subscription") else None),
        sku=(str(normalized["sku"]).strip() if normalized.get("sku") else None),
        environment=(str(normalized["environment"]).strip() if normalized.get("environment") else None),
        urgency=_coerce_urgency(normalized.get("urgency")),
        deployment_type=deployment_type,
        capacity=capacity,
        notes=(str(normalized["notes"]).strip() if normalized.get("notes") else None),
        raw=raw,
    )


def _load_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        # common shapes: {records: [...]}, {deployments: [...]}, {data: [...]}
        for key in ("records", "deployments", "data", "items"):
            if isinstance(data.get(key), list):
                return data[key]
        return [data]
    if isinstance(data, list):
        return data
    return []


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return [dict(r) for r in reader]


def _load_yaml(path: Path) -> list[dict[str, Any]]:
    import yaml  # local import to keep optional footprint light

    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for key in ("records", "deployments", "data", "items"):
            if isinstance(data.get(key), list):
                return data[key]
        return [data]
    if isinstance(data, list):
        return data
    return []


def _load_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        record = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        if any(v not in (None, "") for v in record.values()):
            records.append(record)
    return records


def load_modeliq(path: Path) -> list[ModelIQRecord]:
    """Load a Model IQ export from disk and return normalized records."""
    suffix = path.suffix.lower()
    if suffix == ".json":
        rows: Iterable[dict[str, Any]] = _load_json(path)
    elif suffix == ".csv":
        rows = _load_csv(path)
    elif suffix in (".yaml", ".yml"):
        rows = _load_yaml(path)
    elif suffix == ".xlsx":
        rows = _load_xlsx(path)
    else:
        raise ValueError(f"Unsupported Model IQ file type: {suffix}")

    return [_row_to_record(r) for r in rows if isinstance(r, dict)]
